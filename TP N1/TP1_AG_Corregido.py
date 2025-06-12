import random
import os
import sys
import matplotlib.pyplot as plt # Para graficar (Tener instalado matplotlib)
import pandas as pd # Para guardar en Excel (Tener instalado pandas)
# Utiliza openpyxl para guardar en Excel (Tener instalado openpyxl)

# TP1 - Algoritmos Genéticos

# GENERAL
def limpiar_pantalla():
    if os.name == 'nt':  # Windows
        os.system('cls')
    else:  # Linux o Mac
        os.system('clear')

def aleatorio():
    return random.randint(0, 1)

def completoCromosoma(cantidad):
    cromosoma = [aleatorio() for _ in range(cantidad)]
    return cromosoma

def generarPoblacion(cantidadCromosomas, cantidadGenes):
    poblacion = []
    for i in range(cantidadCromosomas):
        cromosoma = completoCromosoma(cantidadGenes)
        poblacion.append(cromosoma)
    return poblacion

# Fitness
def binarioADecimal(cromosoma):
    decimal = 0
    cantGenes = len(cromosoma)
    for i in range(cantGenes):
        decimal += cromosoma[i] * (2 ** (cantGenes - 1 - i))
    return decimal

def funcionObjetivo(x):
    coef = (2 ** 30) - 1
    return (x / coef) ** 2 

def calculadorEstadisticos(poblacion):
    objetivos = [funcionObjetivo(binarioADecimal(ind)) for ind in poblacion]
    max_objetivos = max(objetivos)
    min_objetivos = min(objetivos)
    mejor_cromosoma = poblacion[objetivos.index(max_objetivos)]
    avg_objetivos = round((sum(objetivos)/len(objetivos)),4)
    #Devuelve tupla con [Max, Min, Avg, Best]
    return [max_objetivos,min_objetivos, avg_objetivos, mejor_cromosoma]  

def calculadorFitness(poblacion):
    fitness = []
    objetivos = []
    sumatoria = 0
    for individuo in poblacion: 
        decimal = binarioADecimal(individuo)
        obj = funcionObjetivo(decimal)
        objetivos.append(obj)
        sumatoria += obj
    for i in range (len(poblacion)): 
        if sumatoria == 0:
            fitness.append(0) #Si la sumatoria es 0, el fitness es 0 para todos.
        else:
            peso = objetivos[i]/sumatoria
            fitness.append(round(peso,5)) #Guardo el peso de cada elemento de la poblacion.
    return fitness

# Crossover
def crossover1Punto(padre, madre):
    puntoCorte = random.randint(1,len(padre)-1)
    h1 = padre[:puntoCorte] + madre[puntoCorte:]
    h2 = madre[:puntoCorte] + padre[puntoCorte:]
    return h1, h2

# Mutacion
def mutacionInvertida(hijo):
    cantGenes = len(hijo)
    if cantGenes < 2:
        return hijo  # No se puede invertir si hay menos de 2 genes
    posInicial = random.randint(0, cantGenes - 2)
    posFinal = random.randint(posInicial + 1, cantGenes - 1)
    segmento = hijo[posInicial:posFinal + 1]
    hijo[posInicial:posFinal + 1] = segmento[::-1]  # Invertir el segmento
    return hijo

# SELECCION

# Ruleta
def seleccionRuleta(poblacion, fitnessValores, cantidad):
    arregloRuleta = []
    seleccionados = []

    for i in range(len(fitnessValores)):
        fitnessValores[i] = fitnessValores[i]*100000
        fitnessValores[i] = int(fitnessValores[i])
        
        for j in range(fitnessValores[i]):
            arregloRuleta.append(poblacion[i]) 
    for i in range (cantidad):
            nro = random.randint(0,99999)
            seleccionados.append(arregloRuleta[nro])
    return seleccionados

# Torneo
def seleccionTorneo(poblacion, fitnessValores, cantidadIndividuos, cantidadCompetidores):
    ganadores = []

    for j in range(cantidadIndividuos):
        competidores = []
        fitness_competidores = []

        for i in range(cantidadCompetidores):
            c=random.randint(0,len(poblacion)-1)
            competidores.append(poblacion[c])
            fitness_competidores.append(fitnessValores[c])

        ganador = competidores[fitness_competidores.index(max(fitness_competidores))]
        ganadores.append(ganador)    
    
    return ganadores

# CICLOS
# Elitismo
def ciclos_con_elitismo(ciclos, prob_crossover, prob_mutacion,cantidadIndividuos, cantidadGenes, metodo_seleccion, cantidadElitismo, cantidadCompetidores=None):
    maximos=[]
    minimos=[]
    promedios=[]
    mejores=[]
    
    pob = generarPoblacion(cantidadIndividuos,cantidadGenes) #Poblacion inicial random
    fit = calculadorFitness(pob)
    rta = calculadorEstadisticos(pob)

    for j in range (ciclos):
        #De la poblacion me quedo con los de elite.
        elitistas = [] 
        fit_ordenados = sorted(fit, reverse=True)
        for i in range(cantidadElitismo):
            indice = fit.index(fit_ordenados[i])
            elitistas.append(pob[indice])
        #elitistas contiene los i mejores individuos de pob

        if metodo_seleccion == 'r':
            pob = seleccionRuleta(pob,fit, cantidadIndividuos-cantidadElitismo) #selecciono una nueva poblacion con los elementos faltantes.
        else:
            pob = seleccionTorneo(pob, fit, cantidadIndividuos-cantidadElitismo, cantidadCompetidores) 
        #REALIZO CROSSOVER Y MUTACION EN LA POBLACION
        
        for i in range (0,len(pob),2):
            padre = pob[i]
            madre = pob[i+1]
            #Si pasa la probablidad de crossover ejecuto el if, me quedo con los hijos.
            if random.random() < prob_crossover :
                hijo1, hijo2 = crossover1Punto(padre,madre)
                pob[i], pob[i+1] = hijo1, hijo2
            
            #Si pasa la probablidad de mutacion me quedo con los mutados (lo hago dos veces, una por elemento)
            if random.random() < prob_mutacion: 
                pob[i] = mutacionInvertida(pob[i])
            if random.random() < prob_mutacion: 
                pob[i+1] = mutacionInvertida(pob[i+1])   
        #FIN CROSSOVER Y MUTACION EN LA POBLACION SELECCIONADA

        #ACTUALMENTE TENGO LA ITERACION 1 EN POBLACION. 
        
        #CALCULO EL FITNESS DE LA NUEVA POBLACION
        pob = pob + elitistas
        fit = calculadorFitness(pob)
        rta = calculadorEstadisticos(pob)

        #GUARDAR VALORES NECESARIOS PARA LA GRAFICA
        maximos.append(rta[0])
        minimos.append(rta[1])
        promedios.append(rta[2])
        mejores.append(rta[3])

        
    return maximos, minimos, promedios, mejores

# Sin elitismo
def ciclos_sin_elitismo(ciclos, prob_crossover, prob_mutacion,cantidadIndividuos, cant_genes, metodo_seleccion, cantidadCompetidores=None):
    maximos=[]
    minimos=[]
    promedios=[]
    mejores=[]
    
    pob = generarPoblacion(cantidadIndividuos,cant_genes)
    fit = calculadorFitness(pob)
    rta = calculadorEstadisticos(pob)
    
    maximos.append(rta[0])
    minimos.append(rta[1])
    promedios.append(rta[2])
    mejores.append(rta[3])

    for j in range (ciclos):
        if metodo_seleccion == 'r':
            pob = seleccionRuleta(pob,fit, cantidadIndividuos) #selecciono una nueva poblacion - trabajo sobre esta.
        else:
            pob = seleccionTorneo(pob, fit, cantidadIndividuos, cantidadCompetidores) 
        #REALIZO CROSSOVER Y MUTACION EN LA POBLACION
        for i in range (0,len(pob),2):
            padre = pob[i]
            madre = pob[i+1]
            #Si pasa la probablidad de crossover ejecuto el if, me quedo con los hijos.
            if random.random() < prob_crossover :
                hijo1, hijo2 = crossover1Punto(padre,madre)
                pob[i], pob[i+1] = hijo1, hijo2
            
            #Si pasa la probablidad de mutacion me quedo con los mutados (lo hago dos veces, una por elemento)
            if random.random() < prob_mutacion: 
                pob[i] = mutacionInvertida(pob[i])
            if random.random() < prob_mutacion: 
                pob[i+1] = mutacionInvertida(pob[i+1])  
        #FIN CROSSOVER Y MUTACION EN LA POBLACION SELECCIONADA

        #ACTUALMENTE TENGO LA ITERACION 1 EN POBLACION.

        #Calculo el fitness de la nueva poblacion
        fit = calculadorFitness(pob)
        rta = calculadorEstadisticos(pob)

        #GUARDAR VALORES NECESARIOS PARA LA GRAFICA
        maximos.append(rta[0])
        minimos.append(rta[1])
        promedios.append(rta[2])
        mejores.append(rta[3])

        
    return maximos, minimos, promedios, mejores

# TABLAS EXCEL
def formatear_hoja_excel(worksheet):
    from openpyxl.styles import Font
    
    # Ajustar ancho de columnas automáticamente
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Poner en negrita la última fila (Promedio)
    bold_font = Font(bold=True)
    last_row = worksheet.max_row
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=last_row, column=col)
        cell.font = bold_font

def crear_tabla_corridas(maximos, minimos, promedios, cantidad_corridas, metodo_seleccion, elitismo_activado):
    metodo_seleccion = metodo_seleccion.strip().lower()
    modo_seleccion = 'RULETA' if metodo_seleccion == 'ruleta' else 'TORNEO'  # Cambiar 'r' por 'ruleta'
    modo_elitismo = 'ELITISTA' if elitismo_activado else 'NO_ELITISTA'
    archivo_excel = f'TP1_{modo_seleccion}_{modo_elitismo}.xlsx'
    nombre_hoja = f'{modo_seleccion}_{modo_elitismo}_{cantidad_corridas}'

    # Crear DataFrame
    df = pd.DataFrame({
        'Corrida': list(range(1, cantidad_corridas + 1)),
        'Máximo Final': maximos,
        'Mínimo Final': minimos,
        'Promedio Final': promedios
    })
    df.loc[len(df)] = [
        'Promedio',  # Corrida sin número
        round(sum(maximos) / cantidad_corridas, 5),
        round(sum(minimos) / cantidad_corridas, 5),
        round(sum(promedios) / cantidad_corridas, 5)
    ]

    try:
        if os.path.exists(archivo_excel):
            with pd.ExcelWriter(archivo_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=nombre_hoja, index=False)
                formatear_hoja_excel(writer.sheets[nombre_hoja])
        else:
            with pd.ExcelWriter(archivo_excel, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=nombre_hoja, index=False)
                formatear_hoja_excel(writer.sheets[nombre_hoja])

        print(f"\nHoja '{nombre_hoja}' guardada en '{archivo_excel}'")

    except PermissionError:
        print(f"\nERROR: El archivo '{archivo_excel}' está abierto y no se puede modificar.")
        print("Cerralo en Excel y volvé a correr el script.")



# GRAFICOS
def generar_grafico_corridas(maximos, minimos, promedios, cantidad_corridas, titulo):
    x = list(range(1, cantidad_corridas + 1))

    fig, axs = plt.subplots(1, 3, figsize=(16, 5))

    axs[0].plot(x, maximos, marker='o', linestyle='-', color='blue', linewidth=0.8)
    axs[0].set_title('Máximos por Corrida')
    axs[0].set_xlabel('Corrida')
    axs[0].set_ylabel('Valor')
    axs[0].grid(True)
    axs[0].set_ylim(0, 1.2)

    axs[1].plot(x, minimos, marker='o', linestyle='-', color='green', linewidth=0.8)
    axs[1].set_title('Mínimos por Corrida')
    axs[1].set_xlabel('Corrida')
    axs[1].grid(True)
    axs[1].set_ylim(0, 1.2)

    axs[2].plot(x, promedios, marker='o', linestyle='-', color='red', linewidth=0.8)
    axs[2].set_title('Promedios por Corrida')
    axs[2].set_xlabel('Corrida')
    axs[2].grid(True)
    axs[2].set_ylim(0, 1.2)

    fig.suptitle(titulo, fontsize=15)
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.savefig(titulo.replace(" ", "_") + '.png')
    plt.show()


# PROGRAMA PRINCIPAL
# Corrida
probCrossover = 0.75
probMutacion = 0.05
cantidadIndividuos = 10
cantidadElitismo = 2
cantidadCompetidores = int(cantidadIndividuos * 0.4)
cantidadCiclosPorCorrida = 20
cantidadGenes = 30
corridas_sin_elitismo = [20, 100, 200]
corridas_con_elitismo = [100]

if len(sys.argv) != 5 or sys.argv[1] != "-s" or sys.argv[3] != "-e":
    print("Uso: python TP1_AG.py -s <seleccion: r-ruleta t-torneo> -e <elitismo: 1-si 0-no>")
    sys.exit(1)
if (sys.argv[2] != "r" and sys.argv[2] != "t") or (int(sys.argv[4]) != 0 and int(sys.argv[4]) != 1):
    print("Error: python TP1_AG.py -s <seleccion: r-ruleta t-torneo> -e <elitismo: 1-si 0-no>")
    sys.exit(1)

# Elección de parámetros
metodoSeleccion = "RULETA" if sys.argv[2] == "r" else "TORNEO"
elitismo_activado = int(sys.argv[4]) == 1
conjuntos_corridas = corridas_con_elitismo if elitismo_activado else corridas_sin_elitismo

# Bucle sobre la cantidad de corridas a ejecutar
for cantidadCorridas in conjuntos_corridas:
    todos_maximos = []
    todos_minimos = []
    todos_promedios = []

    for i in range(cantidadCorridas):
        if elitismo_activado:
            maximos, minimos, promedios, mejores = ciclos_con_elitismo(
                cantidadCiclosPorCorrida,
                probCrossover,
                probMutacion,
                cantidadIndividuos,
                cantidadGenes,
                sys.argv[4],
                cantidadElitismo=cantidadElitismo,
                cantidadCompetidores=cantidadCompetidores
            )
        else:
            maximos, minimos, promedios, mejores = ciclos_sin_elitismo(
                cantidadCiclosPorCorrida,
                probCrossover,
                probMutacion,
                cantidadIndividuos,
                cantidadGenes,
                sys.argv[4],
                cantidadCompetidores=cantidadCompetidores
            )

        # Se almacena el valor final de cada corrida
        todos_maximos.append(maximos[-1])
        todos_minimos.append(minimos[-1])
        todos_promedios.append(promedios[-1])
        print(f"\n--- CORRIDA {i+1} ---")
        print(f"Mejor cromosoma: {mejores[-1]}")
        print(f"Valor decimal: {binarioADecimal(mejores[-1])}")
        print(f"Máximo: {maximos[-1]:.6f}")
        print(f"Mínimo: {minimos[-1]:.6f}")
        print(f"Promedio: {promedios[-1]:.6f}")

    # Generar gráficos y tablas para este conjunto de corridas
    titulo = f"Seleccion {metodoSeleccion} {'ELITISTA' if elitismo_activado else 'NO ELITISTA'} - {cantidadCorridas} corridas, {cantidadCiclosPorCorrida} ciclos cada una"

    generar_grafico_corridas(todos_maximos, todos_minimos, todos_promedios, cantidadCorridas, titulo)
    crear_tabla_corridas(todos_maximos, todos_minimos, todos_promedios, cantidadCorridas, metodoSeleccion, elitismo_activado)

input("\nPresione cualquier tecla para continuar...")
limpiar_pantalla()

# Resumen de todas las corridas
print(f"=== RESUMEN {cantidadCorridas} CORRIDAS ===")
mejor_corrida = todos_maximos.index(max(todos_maximos)) + 1
print(f"Mejor resultado en corrida {mejor_corrida}")
print(f"Máximo global: {max(todos_maximos):.6f}")
print(f"Promedio de máximos: {sum(todos_maximos)/len(todos_maximos):.6f}")